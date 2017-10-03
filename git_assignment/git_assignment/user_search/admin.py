

# from django.http import HttpResponse
from django.contrib import admin
from .models import UserProfile
from git_assignment import settings
from django.http import HttpResponse
import openpyxl

class UserProfileAdmin(admin.ModelAdmin):

    class Meta:
    	model = UserProfile
    """data_record_created_at - time of data record creation"""
    list_display = ("login_name", "login_id", "url", "email", "created_at", "updated_at")
    list_filter = ("created_at", "email", "data_record_created_at")
    actions =  ['export_xlsx']

    def export_xlsx(modeladmin, request, queryset):
        try: 
            from openpyxl.cell import get_column_letter
        except ImportError:
            from openpyxl.utils import get_column_letter
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "MyModel"

        row_num = 0

        columns = [
            (u"ID", 15),
            (u"Name", 70),
            (u"URL", 70),
        ]

        for col_num in xrange(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            
            # set column width
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for obj in queryset:
            row_num += 1
            row = [
                obj.pk,
                obj.login_name,
                obj.url,
            ]
            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
        
        wb.save(response)
        return response

    export_xlsx.short_description = u"Export XLSX"


admin.site.register(UserProfile, UserProfileAdmin)
